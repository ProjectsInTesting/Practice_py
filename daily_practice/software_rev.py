import openpyxl
# 1. Define Excel file path
excel_file = "files/software_versions.xlsx"
try:
    # 2. Load Excel file
    workbook = openpyxl.load_workbook(excel_file)
    # 3. Check if sheet "Versions" exists
    if "Versions" not in workbook.sheetnames:
        print("Error: Sheet 'Versions' not found. Available sheets:", workbook.sheetnames)
        exit()
    sheet = workbook["Versions"]
  
    # 4. Get the last non-empty cell in column A
    latest_version = None
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet[f"A{row}"].value
        if cell_value is not None:  # Find the last non-empty cell
            latest_version = cell_value

    if latest_version is None:
        print("Error: No version found in column A. Add a starting version (e.g., '1.0.0').")
        exit()
    print(f"Current version: {latest_version}")

    # 5. Ask for change type
    change_type = input("Enter the change type (major, minor, patch): ").lower()

    # 6. Split and validate version format (e.g., "23.3.1")
    try:
        major_num, minor_num, patch_num = map(int, latest_version.split("."))
    except ValueError:
        print(f"Error: Invalid version format. Expected 'X.Y.Z', got '{latest_version}'")
        exit()

    # 7. Update version based on change type
    if change_type == "major":
        major_num += 1
        minor_num = 0
        patch_num = 0
    elif change_type == "minor":
        minor_num += 1
        patch_num = 0
    elif change_type == "patch":
        patch_num += 1
    else:
        print("Error: Invalid change type. Use 'major', 'minor', or 'patch'.")
        exit()
    new_version = f"{major_num}.{minor_num}.{patch_num}"
    # 8. Append new version and save
    sheet.append([new_version])
    workbook.save(excel_file)
    print(f"New version saved: {new_version}")

except FileNotFoundError:
    print(f"Error: File '{excel_file}' not found.")
except Exception as e:
    print(f"An unexpected error occurred: {e}")